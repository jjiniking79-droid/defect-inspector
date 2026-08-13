# -*- coding: utf-8 -*-
"""
불량 이미지 자동 판정 프로그램 (Defect Inspector)
------------------------------------------------
- 이미지 폴더/파일 불러오기 (모든 이미지 형식)
- 파일명에서 LOT / GLS / X / Y 좌표 자동 파싱
    예) 64N68001AQ0_64N67002720_64N6700272UC4_TPTN2603_1011_21.829_1203.322_F1II_01.jpg
        -> LOT=64N68001AQ0, GLS=64N67002720, X=21.829, Y=1203.322
- 불량 유형 정의 텍스트 입력/저장 (작성자가 직접 정의, 참고 문서로 저장됨)
- 판정 유형(카테고리) 추가/삭제 관리
- 표 컬럼: 선택 / 이미지 / LOT / GLS / X / Y / AI판정 / 신뢰율(%) / 작업자 판정
- LOT, GLS, X, Y, AI판정, 신뢰율(%), 작업자 판정 필터 기능
- 선택 이미지 일괄 작업자 판정값 입력
- AI 학습하기 / 자동 판정 (신뢰율 70% 미만 행은 파스텔 연핑크 음영)
- MAP 버튼: 선택된 이미지들의 X-Y 좌표를 좌표평면에 타점으로 표시, 점 클릭 시 이미지 확인
- 엑셀로 내보내기 (현재 화면에 보이는 항목 기준)
- 데이터/모델 저장 및 백업

실행: python app.py
배포용 EXE 빌드: build_exe.bat (PyInstaller) 또는 .github/workflows/build.yml (GitHub Actions) 참고
"""

import os
import sys
import threading
import queue
import tempfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from PIL import Image, ImageTk

import data_store
from model_manager import ModelManager

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PyInstaller(onefile)로 실행될 경우, 실행 파일이 위치한 폴더를 작업 폴더로 고정합니다.
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

IMAGE_EXTS = (
    ".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tif", ".tiff",
    ".webp", ".jfif", ".ppm", ".pgm", ".pbm"
)

THUMB_SIZE = (200, 200)
MAP_THUMB_SIZE = (150, 150)

X_RANGE = (0, 1500)
Y_RANGE = (0, 1850)

COLUMNS = ["선택", "이미지", "LOT", "GLS", "X", "Y", "AI판정", "신뢰율(%)", "작업자 판정"]
COL_WIDTHS = [55, 215, 140, 140, 70, 70, 110, 100, 160]
COL_KEYS = ["chk", "thumb", "lot", "gls", "x", "y", "auto", "conf", "worker_combo"]

ROW_BG_NORMAL = "#ffffff"
ROW_BG_LOWCONF = "#ffd9e8"   # 파스텔톤 연핑크
LOW_CONF_THRESHOLD = 70.0    # % 미만이면 음영 처리

ALL_FILTER = "전체"


def parse_filename(path):
    """
    파일명 규칙: LOT_GLS_..._..._..._X_Y_..._...
    예) 64N68001AQ0_64N67002720_64N6700272UC4_TPTN2603_1011_21.829_1203.322_F1II_01
        parts[0]=LOT, parts[1]=GLS, parts[5]=X, parts[6]=Y
    규칙에 맞지 않는 파일명은 LOT/GLS는 "-", X/Y는 None 으로 처리됩니다.
    """
    base = os.path.splitext(os.path.basename(path))[0]
    parts = base.split("_")
    lot = parts[0] if len(parts) > 0 and parts[0] else "-"
    gls = parts[1] if len(parts) > 1 and parts[1] else "-"
    x = y = None
    if len(parts) > 6:
        try:
            x = float(parts[5])
            y = float(parts[6])
        except ValueError:
            x = y = None
    return lot, gls, x, y


def _safe_float(text, default):
    try:
        return float(text)
    except (TypeError, ValueError):
        return default


class ImageRow:
    """표(grid)의 한 행에 대응하는 이미지 레코드"""
    def __init__(self, path):
        self.path = path
        self.lot, self.gls, self.x, self.y = parse_filename(path)
        self.checked = tk.BooleanVar(value=False)
        self.auto_label = tk.StringVar(value="-")
        self.confidence = tk.StringVar(value="-")
        self.worker_label = tk.StringVar(value="")
        self.thumb_img = None  # PhotoImage 참조 유지용
        self.widgets = {}      # 행의 위젯들을 보관 (파괴/재구성/배경색/재배치용)
        self.visible = True    # 필터 결과 표시 여부


class DefectInspectorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("불량 이미지 자동 판정 프로그램")
        self.root.geometry("1420x900")

        self.categories = data_store.load_categories()
        self.rules = data_store.load_rules()
        self.model = ModelManager()
        self.rows = []  # ImageRow 목록

        self.train_queue = queue.Queue()

        self._build_ui()
        self._refresh_all_category_widgets()
        self._refresh_filter_options()
        self._update_status("준비 완료. 모델 학습 여부: " +
                             ("학습됨" if self.model.is_trained() else "미학습"))

    # ------------------------------------------------------------------
    # UI 구성
    # ------------------------------------------------------------------
    def _build_ui(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("LowConf.TCombobox", fieldbackground=ROW_BG_LOWCONF)
        style.configure("Normal.TCombobox", fieldbackground=ROW_BG_NORMAL)

        # ---------- 상단 툴바 1: 불러오기 / 학습 / 판정 / MAP / 엑셀 / 저장 ----------
        toolbar1 = ttk.Frame(self.root, padding=6)
        toolbar1.pack(side="top", fill="x")

        ttk.Button(toolbar1, text="📁 이미지 폴더 불러오기",
                   command=self.load_folder).pack(side="left", padx=3)
        ttk.Button(toolbar1, text="🖼 이미지 파일 불러오기",
                   command=self.load_files).pack(side="left", padx=3)
        ttk.Separator(toolbar1, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(toolbar1, text="🧠 AI 학습하기",
                   command=self.train_model).pack(side="left", padx=3)
        ttk.Button(toolbar1, text="⚡ 자동 판정",
                   command=self.auto_classify).pack(side="left", padx=3)
        ttk.Separator(toolbar1, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(toolbar1, text="🗺 MAP",
                   command=self.open_map).pack(side="left", padx=3)
        ttk.Button(toolbar1, text="📊 엑셀로 내보내기",
                   command=self.export_excel).pack(side="left", padx=3)
        ttk.Separator(toolbar1, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(toolbar1, text="💾 데이터 저장/백업",
                   command=self.backup_data).pack(side="left", padx=3)
        ttk.Button(toolbar1, text="🗑 목록 비우기",
                   command=self.clear_rows).pack(side="left", padx=3)

        self.progress = ttk.Progressbar(toolbar1, mode="determinate", length=160)
        self.progress.pack(side="right", padx=6)
        self.status_var = tk.StringVar(value="")
        ttk.Label(toolbar1, textvariable=self.status_var).pack(side="right", padx=8)

        # ---------- 상단 툴바 2: 판정 유형 관리 + 일괄 판정 ----------
        toolbar2 = ttk.LabelFrame(self.root, text="판정 유형 관리", padding=6)
        toolbar2.pack(side="top", fill="x", padx=6, pady=(0, 4))

        row2a = ttk.Frame(toolbar2)
        row2a.pack(side="top", fill="x")
        ttk.Label(row2a, text="새 유형 추가:").pack(side="left")
        self.new_cat_entry = ttk.Entry(row2a, width=18)
        self.new_cat_entry.pack(side="left", padx=4)
        ttk.Button(row2a, text="추가", command=self.add_category).pack(side="left", padx=4)

        ttk.Separator(row2a, orient="vertical").pack(side="left", fill="y", padx=10)

        ttk.Label(row2a, text="유형 삭제:").pack(side="left")
        self.delete_cat_combo = ttk.Combobox(row2a, values=self.categories,
                                              width=16, state="readonly")
        self.delete_cat_combo.pack(side="left", padx=4)
        ttk.Button(row2a, text="삭제", command=self.delete_category).pack(side="left", padx=4)

        row2b = ttk.Frame(toolbar2)
        row2b.pack(side="top", fill="x", pady=(6, 0))
        ttk.Label(row2b, text="선택한 이미지 일괄 판정값:").pack(side="left")
        self.batch_combo = ttk.Combobox(row2b, values=self.categories,
                                         width=16, state="readonly")
        self.batch_combo.pack(side="left", padx=4)
        ttk.Button(row2b, text="선택 항목에 적용",
                   command=self.apply_batch_label).pack(side="left", padx=4)
        ttk.Button(row2b, text="전체 선택", command=self.select_all).pack(side="left", padx=4)
        ttk.Button(row2b, text="전체 해제", command=self.deselect_all).pack(side="left", padx=4)

        # ---------- 상단 툴바 3: 불량 유형 정의(룰) 텍스트 ----------
        toolbar3 = ttk.LabelFrame(self.root, text="불량 유형 정의 (작성자가 직접 입력 / 판정 참고자료로 저장)", padding=6)
        toolbar3.pack(side="top", fill="x", padx=6, pady=(0, 4))

        rule_top = ttk.Frame(toolbar3)
        rule_top.pack(side="top", fill="x")
        ttk.Label(rule_top, text="유형 선택:").pack(side="left")
        self.rule_combo = ttk.Combobox(rule_top, values=self.categories,
                                        width=16, state="readonly")
        self.rule_combo.pack(side="left", padx=4)
        self.rule_combo.bind("<<ComboboxSelected>>", self.on_rule_category_selected)
        ttk.Button(rule_top, text="정의 저장", command=self.save_rule_text).pack(side="left", padx=4)

        self.rule_text = tk.Text(toolbar3, height=4, wrap="word")
        self.rule_text.pack(side="top", fill="x", pady=(4, 0))

        # ---------- 상단 툴바 4: 필터 ----------
        toolbar4 = ttk.LabelFrame(self.root, text="필터 (LOT / GLS / X / Y / AI판정 / 신뢰율 / 작업자 판정)", padding=6)
        toolbar4.pack(side="top", fill="x", padx=6, pady=(0, 4))

        f1 = ttk.Frame(toolbar4)
        f1.pack(side="top", fill="x")
        ttk.Label(f1, text="LOT:").pack(side="left")
        self.filter_lot = ttk.Combobox(f1, values=[ALL_FILTER], width=14, state="readonly")
        self.filter_lot.set(ALL_FILTER)
        self.filter_lot.pack(side="left", padx=(2, 10))

        ttk.Label(f1, text="GLS:").pack(side="left")
        self.filter_gls = ttk.Combobox(f1, values=[ALL_FILTER], width=14, state="readonly")
        self.filter_gls.set(ALL_FILTER)
        self.filter_gls.pack(side="left", padx=(2, 10))

        ttk.Label(f1, text="X:").pack(side="left")
        self.filter_xmin = ttk.Entry(f1, width=6)
        self.filter_xmin.insert(0, str(X_RANGE[0]))
        self.filter_xmin.pack(side="left", padx=(2, 2))
        ttk.Label(f1, text="~").pack(side="left")
        self.filter_xmax = ttk.Entry(f1, width=6)
        self.filter_xmax.insert(0, str(X_RANGE[1]))
        self.filter_xmax.pack(side="left", padx=(2, 10))

        ttk.Label(f1, text="Y:").pack(side="left")
        self.filter_ymin = ttk.Entry(f1, width=6)
        self.filter_ymin.insert(0, str(Y_RANGE[0]))
        self.filter_ymin.pack(side="left", padx=(2, 2))
        ttk.Label(f1, text="~").pack(side="left")
        self.filter_ymax = ttk.Entry(f1, width=6)
        self.filter_ymax.insert(0, str(Y_RANGE[1]))
        self.filter_ymax.pack(side="left", padx=(2, 2))

        f2 = ttk.Frame(toolbar4)
        f2.pack(side="top", fill="x", pady=(6, 0))
        ttk.Label(f2, text="AI판정:").pack(side="left")
        self.filter_ai = ttk.Combobox(f2, values=[ALL_FILTER] + self.categories, width=14, state="readonly")
        self.filter_ai.set(ALL_FILTER)
        self.filter_ai.pack(side="left", padx=(2, 10))

        ttk.Label(f2, text="신뢰율(%):").pack(side="left")
        self.filter_confmin = ttk.Entry(f2, width=6)
        self.filter_confmin.insert(0, "0")
        self.filter_confmin.pack(side="left", padx=(2, 2))
        ttk.Label(f2, text="~").pack(side="left")
        self.filter_confmax = ttk.Entry(f2, width=6)
        self.filter_confmax.insert(0, "100")
        self.filter_confmax.pack(side="left", padx=(2, 10))

        ttk.Label(f2, text="작업자 판정:").pack(side="left")
        self.filter_worker = ttk.Combobox(f2, values=[ALL_FILTER] + self.categories, width=14, state="readonly")
        self.filter_worker.set(ALL_FILTER)
        self.filter_worker.pack(side="left", padx=(2, 10))

        ttk.Button(f2, text="필터 적용", command=self.apply_filters).pack(side="left", padx=6)
        ttk.Button(f2, text="필터 초기화", command=self.reset_filters).pack(side="left", padx=2)

        # ---------- 안내: 신뢰율 음영 표시 ----------
        legend = ttk.Frame(self.root, padding=(10, 2))
        legend.pack(side="top", fill="x")
        swatch = tk.Label(legend, text="   ", bg=ROW_BG_LOWCONF, relief="solid", borderwidth=1)
        swatch.pack(side="left", padx=(0, 4))
        ttk.Label(legend, text=f"= AI 신뢰율 {LOW_CONF_THRESHOLD:.0f}% 미만 (재검수 권장)").pack(side="left")

        # ---------- 스크롤 가능한 표 영역 (헤더 + 데이터 행이 같은 grid 부모를 공유) ----------
        list_container = ttk.Frame(self.root)
        list_container.pack(side="top", fill="both", expand=True, padx=6, pady=(4, 6))

        self.canvas = tk.Canvas(list_container, borderwidth=0, highlightthickness=0)
        vsb = ttk.Scrollbar(list_container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.table_frame = tk.Frame(self.canvas, bg=ROW_BG_NORMAL)
        self.table_frame_id = self.canvas.create_window((0, 0), window=self.table_frame, anchor="nw")
        self.table_frame.bind("<Configure>",
                               lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>",
                          lambda e: self.canvas.itemconfig(self.table_frame_id, width=e.width))
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        for i, w in enumerate(COL_WIDTHS):
            self.table_frame.grid_columnconfigure(i, minsize=w)

        # 헤더 (0행)
        for i, text in enumerate(COLUMNS):
            tk.Label(self.table_frame, text=text, font=("맑은 고딕", 9, "bold"),
                     bg="#e3e3e3", anchor="center", relief="groove", borderwidth=1
                     ).grid(row=0, column=i, sticky="nsew", padx=0, pady=0, ipady=4)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _update_status(self, text):
        self.status_var.set(text)

    # ------------------------------------------------------------------
    # 카테고리 관리
    # ------------------------------------------------------------------
    def _refresh_all_category_widgets(self):
        self.batch_combo["values"] = self.categories
        self.rule_combo["values"] = self.categories
        self.delete_cat_combo["values"] = self.categories
        self.filter_ai["values"] = [ALL_FILTER] + self.categories
        self.filter_worker["values"] = [ALL_FILTER] + self.categories
        for row in self.rows:
            cb = row.widgets.get("worker_combo")
            if cb is not None:
                cb["values"] = self.categories

    def add_category(self):
        name = self.new_cat_entry.get().strip()
        if not name:
            messagebox.showwarning("알림", "추가할 유형 이름을 입력하세요.")
            return
        if name in self.categories:
            messagebox.showinfo("알림", "이미 존재하는 유형입니다.")
            return
        self.categories.append(name)
        data_store.save_categories(self.categories)
        self.new_cat_entry.delete(0, "end")
        self._refresh_all_category_widgets()
        self._update_status(f"'{name}' 유형이 추가되었습니다.")

    def delete_category(self):
        cat = self.delete_cat_combo.get()
        if not cat:
            messagebox.showwarning("알림", "삭제할 유형을 먼저 선택하세요.")
            return
        if cat not in self.categories:
            return
        if not messagebox.askyesno(
            "삭제 확인",
            f"'{cat}' 유형을 목록에서 삭제하시겠습니까?\n\n"
            f"이미 이 유형으로 저장된 학습 데이터(data/labeled_images/{cat})는 "
            f"삭제되지 않고 그대로 보존됩니다.\n"
            f"단, 새로 판정할 때 더 이상 이 유형을 선택할 수 없게 됩니다."
        ):
            return
        self.categories.remove(cat)
        data_store.save_categories(self.categories)
        if cat in self.rules:
            del self.rules[cat]
            data_store.save_rules(self.rules)
        self._refresh_all_category_widgets()
        self._update_status(f"'{cat}' 유형이 삭제되었습니다.")

    # ------------------------------------------------------------------
    # 불량 유형 정의(룰 텍스트)
    # ------------------------------------------------------------------
    def on_rule_category_selected(self, event=None):
        cat = self.rule_combo.get()
        self.rule_text.delete("1.0", "end")
        self.rule_text.insert("1.0", self.rules.get(cat, ""))

    def save_rule_text(self):
        cat = self.rule_combo.get()
        if not cat:
            messagebox.showwarning("알림", "정의를 저장할 유형을 먼저 선택하세요.")
            return
        self.rules[cat] = self.rule_text.get("1.0", "end").strip()
        data_store.save_rules(self.rules)
        messagebox.showinfo("저장 완료", f"'{cat}' 유형 정의가 저장되었습니다.\n"
                                        f"(이 정의는 판정자 참고 문서로 백업에 함께 저장되며,\n"
                                        f" 실제 자동판정은 AI 학습 데이터를 기반으로 이루어집니다.)")

    # ------------------------------------------------------------------
    # 이미지 불러오기
    # ------------------------------------------------------------------
    def load_folder(self):
        folder = filedialog.askdirectory(title="이미지 폴더 선택")
        if not folder:
            return
        paths = []
        for root_dir, _, files in os.walk(folder):
            for fn in files:
                if fn.lower().endswith(IMAGE_EXTS):
                    paths.append(os.path.join(root_dir, fn))
        if not paths:
            messagebox.showinfo("알림", "선택한 폴더에서 이미지 파일을 찾지 못했습니다.")
            return
        self._add_images(paths)

    def load_files(self):
        filetypes = [
            ("이미지 파일", " ".join(f"*{e}" for e in IMAGE_EXTS)),
            ("모든 파일", "*.*"),
        ]
        paths = filedialog.askopenfilenames(title="이미지 파일 선택", filetypes=filetypes)
        if not paths:
            return
        self._add_images(list(paths))

    def _add_images(self, paths):
        existing = {r.path for r in self.rows}
        added = 0
        for p in paths:
            if p in existing:
                continue
            row = ImageRow(p)
            self._create_row_widgets(row)
            self.rows.append(row)
            added += 1
        self._refresh_filter_options()
        self.apply_filters()
        self._update_status(f"이미지 {added}장 추가됨 (전체 {len(self.rows)}장)")

    def clear_rows(self):
        if not self.rows:
            return
        if not messagebox.askyesno("확인", "불러온 이미지 목록을 모두 비우시겠습니까? (저장된 학습 데이터는 유지됩니다)"):
            return
        for row in self.rows:
            for w in row.widgets.values():
                try:
                    w.destroy()
                except Exception:
                    pass
        self.rows = []
        self._refresh_filter_options()
        self._update_status("목록을 비웠습니다.")

    # ------------------------------------------------------------------
    # 행 위젯 생성 (헤더와 동일한 grid 부모(table_frame) 사용 -> 줄 자동 정렬)
    # 생성 시점에는 grid 배치를 하지 않고, apply_filters()에서 일괄 배치합니다.
    # ------------------------------------------------------------------
    def _create_row_widgets(self, row: ImageRow):
        bg = ROW_BG_NORMAL

        chk = tk.Checkbutton(self.table_frame, variable=row.checked, bg=bg,
                              activebackground=bg, highlightthickness=0)

        try:
            im = Image.open(row.path)
            im.thumbnail(THUMB_SIZE)
            row.thumb_img = ImageTk.PhotoImage(im)
        except Exception:
            row.thumb_img = None

        thumb_lbl = tk.Label(self.table_frame, image=row.thumb_img, bg=bg)

        lot_lbl = tk.Label(self.table_frame, text=row.lot, bg=bg, anchor="center")
        gls_lbl = tk.Label(self.table_frame, text=row.gls, bg=bg, anchor="center")
        x_text = f"{row.x:.0f}" if row.x is not None else "-"
        y_text = f"{row.y:.0f}" if row.y is not None else "-"
        x_lbl = tk.Label(self.table_frame, text=x_text, bg=bg, anchor="center")
        y_lbl = tk.Label(self.table_frame, text=y_text, bg=bg, anchor="center")

        auto_lbl = tk.Label(self.table_frame, textvariable=row.auto_label, bg=bg, anchor="center")
        conf_lbl = tk.Label(self.table_frame, textvariable=row.confidence, bg=bg, anchor="center")

        worker_combo = ttk.Combobox(self.table_frame, textvariable=row.worker_label,
                                     values=self.categories, width=16, state="readonly",
                                     style="Normal.TCombobox")

        row.widgets = {
            "chk": chk, "thumb": thumb_lbl, "lot": lot_lbl, "gls": gls_lbl,
            "x": x_lbl, "y": y_lbl, "auto": auto_lbl, "conf": conf_lbl,
            "worker_combo": worker_combo,
        }

    def _set_row_bg(self, row: ImageRow, low_conf: bool):
        bg = ROW_BG_LOWCONF if low_conf else ROW_BG_NORMAL
        for key in ("chk", "thumb", "lot", "gls", "x", "y", "auto", "conf"):
            w = row.widgets.get(key)
            if w is not None:
                try:
                    w.configure(bg=bg)
                    if key == "chk":
                        w.configure(activebackground=bg)
                except Exception:
                    pass
        combo = row.widgets.get("worker_combo")
        if combo is not None:
            combo.configure(style="LowConf.TCombobox" if low_conf else "Normal.TCombobox")

    # ------------------------------------------------------------------
    # 필터
    # ------------------------------------------------------------------
    def _refresh_filter_options(self):
        lots = sorted({r.lot for r in self.rows if r.lot and r.lot != "-"})
        glss = sorted({r.gls for r in self.rows if r.gls and r.gls != "-"})
        cur_lot = self.filter_lot.get()
        cur_gls = self.filter_gls.get()
        self.filter_lot["values"] = [ALL_FILTER] + lots
        self.filter_gls["values"] = [ALL_FILTER] + glss
        if cur_lot not in self.filter_lot["values"]:
            self.filter_lot.set(ALL_FILTER)
        if cur_gls not in self.filter_gls["values"]:
            self.filter_gls.set(ALL_FILTER)

    def reset_filters(self):
        self.filter_lot.set(ALL_FILTER)
        self.filter_gls.set(ALL_FILTER)
        self.filter_xmin.delete(0, "end"); self.filter_xmin.insert(0, str(X_RANGE[0]))
        self.filter_xmax.delete(0, "end"); self.filter_xmax.insert(0, str(X_RANGE[1]))
        self.filter_ymin.delete(0, "end"); self.filter_ymin.insert(0, str(Y_RANGE[0]))
        self.filter_ymax.delete(0, "end"); self.filter_ymax.insert(0, str(Y_RANGE[1]))
        self.filter_ai.set(ALL_FILTER)
        self.filter_confmin.delete(0, "end"); self.filter_confmin.insert(0, "0")
        self.filter_confmax.delete(0, "end"); self.filter_confmax.insert(0, "100")
        self.filter_worker.set(ALL_FILTER)
        self.apply_filters()

    def apply_filters(self):
        lot = self.filter_lot.get()
        gls = self.filter_gls.get()
        xmin = _safe_float(self.filter_xmin.get(), X_RANGE[0])
        xmax = _safe_float(self.filter_xmax.get(), X_RANGE[1])
        ymin = _safe_float(self.filter_ymin.get(), Y_RANGE[0])
        ymax = _safe_float(self.filter_ymax.get(), Y_RANGE[1])
        ai = self.filter_ai.get()
        confmin = _safe_float(self.filter_confmin.get(), 0)
        confmax = _safe_float(self.filter_confmax.get(), 100)
        worker = self.filter_worker.get()

        x_default = (xmin == X_RANGE[0] and xmax == X_RANGE[1])
        y_default = (ymin == Y_RANGE[0] and ymax == Y_RANGE[1])
        conf_default = (confmin == 0 and confmax == 100)

        for row in self.rows:
            ok = True
            if lot and lot != ALL_FILTER and row.lot != lot:
                ok = False
            if ok and gls and gls != ALL_FILTER and row.gls != gls:
                ok = False
            if ok and row.x is not None and not (xmin <= row.x <= xmax):
                ok = False
            elif ok and row.x is None and not x_default:
                ok = False
            if ok and row.y is not None and not (ymin <= row.y <= ymax):
                ok = False
            elif ok and row.y is None and not y_default:
                ok = False
            if ok and ai and ai != ALL_FILTER and row.auto_label.get() != ai:
                ok = False
            if ok:
                conf_val = _safe_float(row.confidence.get(), None)
                if conf_val is not None:
                    if not (confmin <= conf_val <= confmax):
                        ok = False
                elif not conf_default:
                    ok = False
            if ok and worker and worker != ALL_FILTER and row.worker_label.get() != worker:
                ok = False
            row.visible = ok

        self._relayout_visible_rows()
        shown = sum(1 for r in self.rows if r.visible)
        self._update_status(f"필터 적용됨: {shown} / {len(self.rows)} 장 표시 중")

    def _relayout_visible_rows(self):
        r = 1  # 0행은 헤더
        for row in self.rows:
            if row.visible:
                for col, key in enumerate(COL_KEYS):
                    w = row.widgets[key]
                    padx = 4 if key == "thumb" else (6 if key in ("lot", "gls", "worker_combo") else 2)
                    pady = 4 if key in ("thumb", "lot", "gls", "x", "y", "auto", "conf", "worker_combo") else 1
                    w.grid(row=r, column=col, sticky="nsew", padx=padx, pady=pady)
                self.table_frame.grid_rowconfigure(r, minsize=THUMB_SIZE[1] + 8)
                r += 1
            else:
                for w in row.widgets.values():
                    w.grid_remove()

    # ------------------------------------------------------------------
    # 선택/일괄 판정
    # ------------------------------------------------------------------
    def select_all(self):
        for r in self.rows:
            if r.visible:
                r.checked.set(True)

    def deselect_all(self):
        for r in self.rows:
            r.checked.set(False)

    def apply_batch_label(self):
        cat = self.batch_combo.get()
        if not cat:
            messagebox.showwarning("알림", "적용할 판정 유형을 먼저 선택하세요.")
            return
        checked = [r for r in self.rows if r.checked.get()]
        if not checked:
            messagebox.showwarning("알림", "선택(체크)된 이미지가 없습니다.")
            return
        for r in checked:
            r.worker_label.set(cat)
        self._update_status(f"선택된 {len(checked)}장에 '{cat}' 판정값을 적용했습니다.")

    # ------------------------------------------------------------------
    # AI 학습
    # ------------------------------------------------------------------
    def train_model(self):
        session_labeled = [(r.path, r.worker_label.get()) for r in self.rows
                            if r.worker_label.get().strip()]

        if not session_labeled:
            dataset = data_store.load_dataset()
            if len(dataset) < 2:
                messagebox.showwarning(
                    "알림",
                    "학습할 데이터가 없습니다.\n"
                    "먼저 이미지에 '작업자 판정값'을 입력한 뒤 학습하기를 눌러주세요."
                )
                return
            combined = dataset
        else:
            for path, label in session_labeled:
                try:
                    data_store.add_labeled_image(path, label)
                except Exception as e:
                    print("labeled image copy failed:", e)
            dataset = data_store.load_dataset()
            merged = {}
            for path, label in dataset:
                merged[path] = label
            for path, label in session_labeled:
                merged[path] = label
            combined = list(merged.items())

        if len(set(l for _, l in combined)) < 2:
            messagebox.showwarning("알림", "학습을 위해서는 서로 다른 판정 유형이 최소 2종류 이상 필요합니다.")
            return

        self._set_controls_enabled(False)
        self.progress.configure(mode="determinate", maximum=len(combined), value=0)
        self._update_status(f"AI 학습 중... (총 {len(combined)}장)")

        def worker():
            try:
                paths = [p for p, _ in combined]
                labels = [l for _, l in combined]

                def progress_cb(i, total):
                    self.train_queue.put(("progress", i, total))

                acc = self.model.train(paths, labels, progress_cb=progress_cb)
                self.train_queue.put(("done", acc, len(combined)))
            except Exception as e:
                self.train_queue.put(("error", str(e)))

        threading.Thread(target=worker, daemon=True).start()
        self.root.after(100, self._poll_train_queue)

    def _poll_train_queue(self):
        try:
            while True:
                msg = self.train_queue.get_nowait()
                if msg[0] == "progress":
                    _, i, total = msg
                    self.progress.configure(value=i)
                    self._update_status(f"특징 추출 중... {i}/{total}")
                elif msg[0] == "done":
                    _, acc, n = msg
                    self._set_controls_enabled(True)
                    self.progress.configure(value=0)
                    acc_txt = f"{acc*100:.1f}%" if acc is not None else "N/A (데이터 부족으로 검증 생략)"
                    self._update_status(f"학습 완료. 데이터 {n}장, 교차검증 정확도: {acc_txt}")
                    messagebox.showinfo("학습 완료",
                                         f"총 {n}장의 데이터로 학습을 완료했습니다.\n"
                                         f"교차검증 정확도: {acc_txt}")
                    return
                elif msg[0] == "error":
                    self._set_controls_enabled(True)
                    self.progress.configure(value=0)
                    messagebox.showerror("학습 오류", msg[1])
                    self._update_status("학습 중 오류가 발생했습니다.")
                    return
        except queue.Empty:
            pass
        self.root.after(100, self._poll_train_queue)

    def _set_controls_enabled(self, enabled):
        state = "normal" if enabled else "disabled"
        def _walk(widget):
            for child in widget.winfo_children():
                if isinstance(child, (ttk.Button,)):
                    child.configure(state=state)
                _walk(child)
        _walk(self.root)

    # ------------------------------------------------------------------
    # 자동 판정
    # ------------------------------------------------------------------
    def auto_classify(self):
        if not self.model.is_trained():
            messagebox.showwarning("알림", "먼저 'AI 학습하기'로 모델을 학습시켜야 자동 판정을 사용할 수 있습니다.")
            return
        if not self.rows:
            messagebox.showwarning("알림", "판정할 이미지가 없습니다. 먼저 이미지를 불러오세요.")
            return

        self._set_controls_enabled(False)
        self.progress.configure(mode="determinate", maximum=len(self.rows), value=0)
        self._update_status("자동 판정 진행 중...")

        def worker():
            for i, row in enumerate(self.rows):
                try:
                    label, conf = self.model.predict(row.path)
                except Exception:
                    label, conf = "오류", 0.0
                self.train_queue.put(("predict_row", row, label, conf, i + 1, len(self.rows)))
            self.train_queue.put(("predict_done",))

        threading.Thread(target=worker, daemon=True).start()
        self.root.after(100, self._poll_predict_queue)

    def _poll_predict_queue(self):
        try:
            while True:
                msg = self.train_queue.get_nowait()
                if msg[0] == "predict_row":
                    _, row, label, conf, i, total = msg
                    row.auto_label.set(label if label else "-")
                    conf_pct = conf * 100 if label else 0.0
                    row.confidence.set(f"{conf_pct:.1f}" if label else "-")
                    self._set_row_bg(row, low_conf=bool(label) and conf_pct < LOW_CONF_THRESHOLD)
                    self.progress.configure(value=i)
                    self._update_status(f"자동 판정 중... {i}/{total}")
                elif msg[0] == "predict_done":
                    self._set_controls_enabled(True)
                    self.progress.configure(value=0)
                    self._update_status("자동 판정이 완료되었습니다.")
                    self.apply_filters()
                    return
        except queue.Empty:
            pass
        self.root.after(100, self._poll_predict_queue)

    # ------------------------------------------------------------------
    # MAP
    # ------------------------------------------------------------------
    def open_map(self):
        selected = [r for r in self.rows if r.checked.get()]
        if not selected:
            messagebox.showwarning("알림", "MAP에 표시할 이미지를 먼저 체크박스로 선택하세요.")
            return
        selected = [r for r in selected if r.x is not None and r.y is not None]
        if not selected:
            messagebox.showwarning("알림", "선택된 이미지 중 X/Y 좌표를 인식할 수 있는 이미지가 없습니다.\n"
                                          "(파일명 규칙을 확인해주세요)")
            return

        popup = tk.Toplevel(self.root)
        popup.title(f"MAP - 선택 이미지 {len(selected)}개")
        popup.geometry("1150x760")

        main_frame = ttk.Frame(popup)
        main_frame.pack(fill="both", expand=True, padx=8, pady=8)

        map_frame = ttk.LabelFrame(
            main_frame, text="X-Y 좌표 맵 (점 클릭 또는 마우스 드래그로 영역 선택 - 여러 점 선택 가능)")
        map_frame.pack(side="left", fill="y", padx=(0, 8))

        CANVAS_W, CANVAS_H = 690, 560
        MARGIN = 34
        plot_w = CANVAS_W - 2 * MARGIN
        plot_h = CANVAS_H - 2 * MARGIN

        map_canvas = tk.Canvas(map_frame, width=CANVAS_W, height=CANVAS_H, bg="white",
                                highlightthickness=0)
        map_canvas.pack(padx=6, pady=6)

        map_canvas.create_rectangle(MARGIN, MARGIN, MARGIN + plot_w, MARGIN + plot_h, outline="#888")
        # X축(세로): 아래쪽이 0, 위쪽이 1500
        map_canvas.create_text(MARGIN - 4, MARGIN, text=f"X:{X_RANGE[1]}", anchor="e", font=("맑은 고딕", 8))
        map_canvas.create_text(MARGIN - 4, MARGIN + plot_h, text=f"X:{X_RANGE[0]}", anchor="e",
                                font=("맑은 고딕", 8))
        # Y축(가로): 왼쪽이 0, 오른쪽이 1850
        map_canvas.create_text(MARGIN, MARGIN + plot_h + 14, text=f"Y:{Y_RANGE[0]}", anchor="w",
                                font=("맑은 고딕", 8))
        map_canvas.create_text(MARGIN + plot_w, MARGIN + plot_h + 14, text=f"Y:{Y_RANGE[1]}", anchor="e",
                                font=("맑은 고딕", 8))

        def to_canvas_xy(x, y):
            # 좌측 하단이 (X=0, Y=0), 좌측 상단이 X=1500 이 되도록
            # X는 세로축(아래->위로 증가), Y는 가로축(왼쪽->오른쪽으로 증가)
            cx = MARGIN + (y / Y_RANGE[1]) * plot_w
            cy = MARGIN + plot_h - (x / X_RANGE[1]) * plot_h
            return cx, cy

        dot_refs = {}
        for row in selected:
            cx, cy = to_canvas_xy(row.x, row.y)
            dot_id = map_canvas.create_oval(cx - 5, cy - 5, cx + 5, cy + 5,
                                             fill="#4a90d9", outline="#20406e", width=1)
            dot_refs[dot_id] = row

        detail_frame = ttk.LabelFrame(main_frame, text="선택된 좌표 이미지")
        detail_frame.pack(side="left", fill="both", expand=True)

        detail_canvas = tk.Canvas(detail_frame, borderwidth=0, highlightthickness=0)
        detail_vsb = ttk.Scrollbar(detail_frame, orient="vertical", command=detail_canvas.yview)
        detail_canvas.configure(yscrollcommand=detail_vsb.set)
        detail_vsb.pack(side="right", fill="y")
        detail_canvas.pack(side="left", fill="both", expand=True)

        detail_inner = ttk.Frame(detail_canvas)
        detail_window = detail_canvas.create_window((0, 0), window=detail_inner, anchor="nw")
        detail_inner.bind("<Configure>",
                           lambda e: detail_canvas.configure(scrollregion=detail_canvas.bbox("all")))
        detail_canvas.bind("<Configure>",
                            lambda e: detail_canvas.itemconfig(detail_window, width=e.width))

        popup._map_photo_refs = []
        map_selected_set = set()

        def refresh_detail_panel():
            for w in detail_inner.winfo_children():
                w.destroy()
            popup._map_photo_refs.clear()
            if not map_selected_set:
                ttk.Label(detail_inner, text="맵에서 점을 클릭하면 해당 이미지가 여기에 표시됩니다.",
                          padding=10).pack(anchor="w")
                return
            for row in map_selected_set:
                item = ttk.Frame(detail_inner, padding=6, relief="groove")
                item.pack(side="top", fill="x", pady=3, padx=3)
                try:
                    im = Image.open(row.path)
                    im.thumbnail(MAP_THUMB_SIZE)
                    photo = ImageTk.PhotoImage(im)
                    popup._map_photo_refs.append(photo)
                    tk.Label(item, image=photo).pack(side="left", padx=(0, 10))
                except Exception:
                    ttk.Label(item, text="[이미지 로드 실패]").pack(side="left", padx=(0, 10))
                info = ttk.Frame(item)
                info.pack(side="left", anchor="n")
                ttk.Label(info, text=f"X: {row.x:.0f}   Y: {row.y:.0f}",
                          font=("맑은 고딕", 10, "bold")).pack(anchor="w")
                ttk.Label(info, text=f"LOT: {row.lot}").pack(anchor="w")
                ttk.Label(info, text=f"GLS: {row.gls}").pack(anchor="w")
                ttk.Label(info, text=f"AI판정: {row.auto_label.get()}   신뢰율: {row.confidence.get()}%").pack(anchor="w")
                ttk.Label(info, text=f"작업자 판정: {row.worker_label.get() or '-'}").pack(anchor="w")

        def select_dot(dot_id, row, exclusive_toggle=True):
            if exclusive_toggle and row in map_selected_set:
                map_selected_set.remove(row)
                map_canvas.itemconfig(dot_id, fill="#4a90d9", outline="#20406e")
            else:
                map_selected_set.add(row)
                map_canvas.itemconfig(dot_id, fill="#ff5a8c", outline="#a10047")

        drag_state = {"start": None, "rect_id": None}
        DRAG_THRESHOLD = 4  # 이 이상 움직이면 클릭이 아닌 드래그(영역 선택)로 처리

        def on_map_press(event):
            drag_state["start"] = (event.x, event.y)
            if drag_state["rect_id"] is None:
                drag_state["rect_id"] = map_canvas.create_rectangle(
                    event.x, event.y, event.x, event.y,
                    outline="#ff5a8c", dash=(4, 2))
            else:
                map_canvas.coords(drag_state["rect_id"], event.x, event.y, event.x, event.y)

        def on_map_drag(event):
            if drag_state["start"] is None or drag_state["rect_id"] is None:
                return
            sx, sy = drag_state["start"]
            map_canvas.coords(drag_state["rect_id"], sx, sy, event.x, event.y)

        def on_map_release(event):
            if drag_state["start"] is None:
                return
            sx, sy = drag_state["start"]
            dx, dy = abs(event.x - sx), abs(event.y - sy)

            if drag_state["rect_id"] is not None:
                map_canvas.delete(drag_state["rect_id"])
                drag_state["rect_id"] = None

            changed = False
            if dx < DRAG_THRESHOLD and dy < DRAG_THRESHOLD:
                # 단순 클릭: 가장 가까운 점 하나를 토글 선택
                closest = map_canvas.find_closest(event.x, event.y)
                if closest:
                    dot_id = closest[0]
                    row = dot_refs.get(dot_id)
                    if row is not None:
                        coords = map_canvas.coords(dot_id)
                        if coords:
                            ccx, ccy = (coords[0] + coords[2]) / 2, (coords[1] + coords[3]) / 2
                            if ((event.x - ccx) ** 2 + (event.y - ccy) ** 2) ** 0.5 <= 10:
                                select_dot(dot_id, row, exclusive_toggle=True)
                                changed = True
            else:
                # 드래그: 사각형 영역 안의 모든 점을 선택에 추가
                x0, x1 = sorted((sx, event.x))
                y0, y1 = sorted((sy, event.y))
                for dot_id, row in dot_refs.items():
                    coords = map_canvas.coords(dot_id)
                    if not coords:
                        continue
                    ccx, ccy = (coords[0] + coords[2]) / 2, (coords[1] + coords[3]) / 2
                    if x0 <= ccx <= x1 and y0 <= ccy <= y1 and row not in map_selected_set:
                        select_dot(dot_id, row, exclusive_toggle=False)
                        changed = True

            drag_state["start"] = None
            if changed:
                refresh_detail_panel()

        map_canvas.bind("<ButtonPress-1>", on_map_press)
        map_canvas.bind("<B1-Motion>", on_map_drag)
        map_canvas.bind("<ButtonRelease-1>", on_map_release)
        refresh_detail_panel()

    # ------------------------------------------------------------------
    # 엑셀 내보내기
    # ------------------------------------------------------------------
    def export_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("오류", "엑셀 내보내기 기능을 사용하려면 'openpyxl' 라이브러리가 필요합니다.\n"
                                        "requirements.txt에 openpyxl을 추가한 뒤 다시 빌드해주세요.")
            return
        visible_rows = [r for r in self.rows if r.visible]
        if not visible_rows:
            messagebox.showwarning("알림", "내보낼 항목이 없습니다. (필터 결과가 비어있습니다)")
            return

        dest = filedialog.asksaveasfilename(
            title="엑셀로 저장", defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")], initialfile="판정결과.xlsx")
        if not dest:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "판정결과"
            ws.append(["선택", "이미지", "LOT", "GLS", "X", "Y", "AI판정", "신뢰율(%)", "작업자 판정"])

            widths = {"A": 6, "B": 16, "C": 16, "D": 16, "E": 8, "F": 8, "G": 12, "H": 12, "I": 14}
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

            tmp_dir = tempfile.mkdtemp(prefix="defect_inspector_xlsx_")
            for i, row in enumerate(visible_rows, start=2):
                ws.cell(row=i, column=1, value="O" if row.checked.get() else "")
                ws.cell(row=i, column=3, value=row.lot)
                ws.cell(row=i, column=4, value=row.gls)
                ws.cell(row=i, column=5, value=round(row.x) if row.x is not None else "")
                ws.cell(row=i, column=6, value=round(row.y) if row.y is not None else "")
                ws.cell(row=i, column=7, value=row.auto_label.get())
                ws.cell(row=i, column=8, value=row.confidence.get())
                ws.cell(row=i, column=9, value=row.worker_label.get())
                ws.row_dimensions[i].height = 66

                try:
                    thumb_path = os.path.join(tmp_dir, f"thumb_{i}.png")
                    im = Image.open(row.path).convert("RGB")
                    im.thumbnail((85, 85))
                    im.save(thumb_path)
                    xlimg = XLImage(thumb_path)
                    ws.add_image(xlimg, f"B{i}")
                except Exception:
                    pass

            wb.save(dest)
            messagebox.showinfo("완료", f"엑셀 파일로 저장되었습니다.\n\n{dest}")
            self._update_status(f"엑셀 내보내기 완료: {dest}")
        except Exception as e:
            messagebox.showerror("엑셀 저장 오류", str(e))

    # ------------------------------------------------------------------
    # 저장/백업
    # ------------------------------------------------------------------
    def backup_data(self):
        dest = filedialog.askdirectory(title="백업할 위치 선택")
        if not dest:
            return
        try:
            target = data_store.backup_all(dest)
            messagebox.showinfo("백업 완료", f"학습 모델 및 데이터가 백업되었습니다.\n\n{target}")
            self._update_status(f"백업 완료: {target}")
        except Exception as e:
            messagebox.showerror("백업 오류", str(e))


def main():
    root = tk.Tk()
    app = DefectInspectorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
